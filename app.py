import telebot, json, random
from __sdk_strop_rozpocet_modul import rozpocet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side


TOKEN='6921653160:AAE_FM06W--mMs7TxabUj6quG1Il9VliNGs'
bot = telebot.TeleBot(TOKEN)

def is_float(co):
		try: float(co); return True
		except: return False	


data = {}

def cislo(s):
	str_cislo = ''
	tecka = 0
	
	for a in s.strip():
		if a.isdigit():
			str_cislo += a
		elif a == '.' and tecka == 0:
			str_cislo += a
			tecka = 1
		elif a == ',' and tecka == 0:
			str_cislo += '.'
			tecka = 1
	try:
		c = float(str_cislo)
	except: c = 0
	return c


def start(_id, _name, text):
	odpoved = '''
	**************************
	    VÍTÁM U VLAD_BOTa
	Na výběr máte komandy:

    👉 /SDK_STROP
    
    👉 /ANEKDOT
    
    👉 /EXIT ❌

	**************************'''
	return odpoved


def sdk_strop(_id, _name, text):
	global data


	if text == '/EXIT':
		data[_id]['data'] = []
		return ' !!! SDK_STROP DATA VYMAZÁNA !!!'


	for d in data[_id]['data']:
		if d[1] == 0:
			neni = data[_id]['data'].index(d)
			break
		else: neni = 'OK'
		
	if neni == 'OK':
		vysledek = '<<< ??? >>>\n /EXIT ❌'
		data[_id]['data'] = []
		data[_id]['method'] = start
		return vysledek

	elif text == '/SDK_STROP':
		return data[_id]['data'][neni][2]

	else:
		if neni == 0:
			if cislo(text) != 0:
				data[_id]['data'][neni][1] = cislo(text)
				print(f'________Uveden cislo: {text} >>> {cislo(text)}')
				return data[_id]['data'][neni+1][2]
			else: return '!!! Číslo špatně\n'+data[_id]['data'][neni][2]

		elif neni == 1:
			if cislo(text) != 0:
				data[_id]['data'][neni][1] = cislo(text)
				print(f'_______Uveden cislo: {text} >>> {cislo(text)}')
				return data[_id]['data'][neni+1][2]
			else: return '!!! Číslo špatně\n'+data[_id]['data'][neni][2]

		elif neni == 2:
			if text == '/3_METROVY':
				data[_id]['data'][neni][1] = '3'
				return data[_id]['data'][neni+1][2]
			elif text == '/4_METROVY':
				data[_id]['data'][neni][1] = '4'
				return data[_id]['data'][neni+1][2]
			else: return '!!! Cekam na jina data.\n'+data[_id]['data'][neni][2]

		elif neni == 3:
			if text == '/JEDEN_ZAKLOP':
				data[_id]['data'][neni][1] = '1'
				return data[_id]['data'][neni+1][2]
			elif text == '/DVA_ZAKLOPY':
				data[_id]['data'][neni][1] = '2'
				return data[_id]['data'][neni+1][2]
			else: return '!!! Čekám na jiná data.\n'+data[_id]['data'][neni][2]

		elif neni == 4:
			if text == '/PRIMY_ZAVES':
				data[_id]['data'][neni][1] = 'primy'
				return data[_id]['data'][neni+1][2]
			elif text == '/DRAT':
				data[_id]['data'][neni][1] = 'drat'
				return data[_id]['data'][neni+1][2]
			else: return '!!! Čekám na jiná data.\n'+data[_id]['data'][neni][2]

		elif neni == 5:
			if text == '/VATA_50mm':
				data[_id]['data'][neni][1] = '50'
				return data[_id]['data'][neni+1][2]
			elif text == '/VATA_60mm':
				data[_id]['data'][neni][1] = '60'
				return data[_id]['data'][neni+1][2]
			elif text == '/VATA_80mm':
				data[_id]['data'][neni][1] = '80'
				return data[_id]['data'][neni+1][2]
			elif text == '/VATA_100mm':
				data[_id]['data'][neni][1] = '100'
				return data[_id]['data'][neni+1][2]
			elif text == '/VATA_NE':
				data[_id]['data'][neni][1] = 'ne'
				return data[_id]['data'][neni+1][2]
			else: return '!!! Čekám na jiná data.\n'+data[_id]['data'][neni][2]

		elif neni == 6:
			if text == '/FOLIJA_ANO':
				data[_id]['data'][neni][1] = 'ano'
			elif text == '/FOLIJA_NE':
				data[_id]['data'][neni][1] = 'ne'
			else: return '!!! Cekam na jina data.\n'+data[_id]['data'][neni][2]

			ar = rozpocet(data[_id]['data'])
			ar.insert(0, ['Material', 'Pocet', 'Cena'])
			file_name = '__TEMP__'+str(_id)+'_'+_name+'.xlsx'
			
			wb = Workbook()
			ws = wb.active
			ws.title = 'SDK_STROP'
			for row in ar:
				ws.append(row)
			
			ws.column_dimensions['A'].width = 40
			ws.column_dimensions['B'].width = 20
			ws.column_dimensions['C'].width = 20

			okraje = Border(left=Side(style='medium'), 
		                     right=Side(style='medium'), 
		                     top=Side(style='medium'), 
		                     bottom=Side(style='medium'))


			for cell in ws['A']:
				if cell.value:
					cell.alignment = Alignment(wrap_text=True)
					cell.font = Font(italic=True, size=16)
					cell.border = okraje
					posledni = cell.row 
			for cell in ws['B']:
				if cell.value:
					cell.alignment = Alignment(horizontal="center", vertical="center")
					cell.font = Font(size=16)
					cell.border = okraje
			for cell in ws['C']:
				if cell.value:
					cell.alignment = Alignment(horizontal="center", vertical="center")
					cell.font = Font(size=16)
					cell.border = okraje
			for cell in ws[1]:
				if cell.value:
					cell.alignment = Alignment(horizontal="center", vertical="center")
					cell.font = Font(name='Times New Roman', size=22, color='FF0000', bold=True)

			for cell in ws[posledni]:
				if cell.value:
					cell.alignment = Alignment(horizontal="center", vertical="center")
					cell.font = Font(size=22, bold=True)	

			wb.save(file_name)

			bot.send_document(_id, open(file_name, 'rb'))
			print(f'Odeslan soubor: {file_name}')

			data[_id]['data'] = []
			data[_id]['method'] = start

			return '👆 Vytvořen soubor s tabulko\n Kliknete na soubor pro otevreni.\n\n👉 /EXIT ❌'

def anekdot(_id, _name, text):
	global data
	try:
		f = open('vtipy.json', 'r', encoding='utf-8')
		vtipy = json.load(f)
		f.close()
		vysledek = random.choice(vtipy)
	except: vysledek = "Anekdot nenalezen :)"
	data[_id]['method'] = start
	return vysledek  + '\n\n 👉 /ANEKDOT\n\n👉 /EXIT ❌'

def trasser(_id, _name, text):
	global data
	if _id not in data:
		data[_id] = {'method': start}

	if text == '/EXIT':
		odpoved = data[_id]['method'](_id, _name, text)
		if data[_id]['method'] != start:
			data[_id]['method'] = start
			odpoved += data[_id]['method'](_id, _name, text)
		return odpoved

	elif text == '/SDK_STROP':
		with open('__SABLONY_OTAZEK.json') as file:
			json_temp = json.load(file)
		data[_id]['method'] = sdk_strop
		data[_id]['data'] = json_temp['sdk_strop']
		odpoved = data[_id]['method'](_id, _name, text)
		return odpoved
	
	elif text == '/ANEKDOT':
		data[_id]['method'] =  anekdot
		odpoved = data[_id]['method'](_id, _name, text)
		return odpoved

	else:
		odpoved = data[_id]['method'](_id, _name, text)
		return odpoved


@bot.message_handler(content_types=['text'])
def text_message(message):
	_text = message.text
	_name = str(message.from_user.first_name)+'_'+str(message.from_user.last_name)
	_id = message.chat.id
	print(str(message.chat.id)+'. '+_name+'. Prijato: '+_text)
	odpoved = trasser(_id, _name, _text)
		
	bot.send_message(_id, odpoved)

print ('_______________Start BOT !!!')
bot.infinity_polling()
